from flask import Flask, request, jsonify, send_from_directory
from flask import render_template
from collections import OrderedDict
import pandas as pd
import json
import os
import datetime
import requests
from openpyxl import load_workbook
from copy import copy
from openpyxl.cell.cell import MergedCell
import qrcode
from openpyxl.drawing.image import Image as ExcelImage
import tempfile
from openpyxl.styles import Font
from openpyxl.drawing.image import Image as OpenPyxlImage
import pdfkit
#import win32com.client as win32
#import pythoncom
import math


app = Flask(__name__)
app.config['UPLOAD_FOLDER'] = 'uploads'
os.makedirs(app.config['UPLOAD_FOLDER'], exist_ok=True)

CONFIG = {
    "sandbox": {
        "RECORDS_FILE": "records_sandbox.json",
        "API_TOKEN": "20352731-ac43-3e3b-99a5-f0a592f7ccfc",
        "API_URL": "https://gw.fbr.gov.pk/di_data/v1/di/postinvoicedata_sb"
    },
    "production": {
        "RECORDS_FILE": "records_production.json",
        "API_TOKEN": "PRODUCTION_TOKEN",
        "API_URL": "https://production-url.example.com"
    }
}

def get_env():
    env = request.args.get('env') or request.headers.get('X-ERP-ENV') or 'sandbox'
    return env if env in CONFIG else 'sandbox'

def get_records_file(env):
    return CONFIG[env]['RECORDS_FILE']

def get_api_token(env):
    return CONFIG[env]['API_TOKEN']

def get_api_url(env):
    return CONFIG[env]['API_URL']

def load_records(env):
    file = get_records_file(env)
    if os.path.exists(file):
        with open(file, 'r') as f:
            try:
                return json.load(f)
            except Exception:
                return []
    return []

def save_records(env, records):
    file = get_records_file(env)
    with open(file, 'w') as f:
        json.dump(records, f, indent=2)

# Store last uploaded file and last JSON per environment
last_uploaded_file = {}
last_json_data = {}

#  Upload Excel File
@app.route('/upload-excel', methods=['POST'])
def upload_excel():
    env = get_env()
    file = request.files.get('file')
    if not file or not file.filename.endswith(('.xlsx', '.xls')):
        return jsonify({'error': 'Invalid file format'}), 400
    filepath = os.path.join(app.config['UPLOAD_FOLDER'], f"{env}_{file.filename}")
    file.save(filepath)
    last_uploaded_file[env] = filepath
    return jsonify({'message': 'File uploaded successfully'})




# Get JSON Data
@app.route('/get-json', methods=['GET'])
def get_json():
    env = get_env()
    if env not in last_uploaded_file or not os.path.exists(last_uploaded_file[env]):
        print("ERROR: File not found or not uploaded. Env =", env)
        print("last_uploaded_file =", last_uploaded_file)
        return jsonify({'error': 'No file uploaded'}), 400

    def safe(val, default=""):
        return default if pd.isna(val) or val in [None, ""] else val

    df = pd.read_excel(last_uploaded_file[env], header=None)

    # Parse sectioned key-value pairs
    section_data = {}
    product_start_index = None

    for i, row in df.iterrows():
        key = str(row[0]).strip() if pd.notna(row[0]) else ''
        val = row[1] if len(row) > 1 else None

        if key.lower() == "hscode":
            product_start_index = i
            break

        if key and not any(key.startswith(s) for s in ["1)", "2)", "3)", "4)"]):
            section_data[key] = safe(val, "")

    if product_start_index is None:
        print("ERROR: No product section found. Section data:", section_data)
        print("File path:", last_uploaded_file[env])
        return jsonify({'error': 'No product section found'}), 400

    product_df = pd.read_excel(last_uploaded_file[env], skiprows=product_start_index)

    items = []
    for _, row in product_df.iterrows():
        try:
            rate_raw = safe(row.get("rate", ""), "")
            rate = str(rate_raw).strip() if isinstance(rate_raw, str) else f"{int(float(rate_raw) * 100)}%"

            hs_code_raw = safe(row.get("hsCode", ""))
            hs_code = (
                "{:.4f}".format(float(hs_code_raw))
                if isinstance(hs_code_raw, (int, float)) and not pd.isna(hs_code_raw)
                else str(hs_code_raw).strip()
            )

            item = OrderedDict([
                ("hsCode", hs_code),
                ("productDescription", safe(row.get("productDescription"))),
                ("rate", rate),
                ("uoM", safe(row.get("uoM"))),
                ("quantity", int(safe(row.get("quantity"), 0))),
                ("totalValues", float(safe(row.get("totalValues"), 0))),
                ("valueSalesExcludingST", float(safe(row.get("valueSalesExcludingST"), 0))),
                ("fixedNotifiedValueOrRetailPrice", float(safe(row.get("fixedNotifiedValueOrRetailPrice"), 0))),
                ("salesTaxApplicable", float(safe(row.get("salesTaxApplicable"), 0))),
                ("salesTaxWithheldAtSource", float(safe(row.get("salesTaxWithheldAtSource"), 0))),
                ("extraTax", str(safe(row.get("extraTax")))),
                ("furtherTax", float(safe(row.get("furtherTax"), 0))),
                ("sroScheduleNo", str(safe(row.get("sroScheduleNo")))),
                ("fedPayable", float(safe(row.get("fedPayable"), 0))),
                ("discount", float(safe(row.get("discount"), 0))),
                ("saleType", str(safe(row.get("saleType")))),
                ("sroItemSerialNo", str(safe(row.get("sroItemSerialNo"))))
            ])
            items.append(item)
        except Exception as e:
            return jsonify({"error": f"Error parsing row: {row.to_dict()} — {str(e)}"}), 400

    raw_date = section_data.get("invoiceDate", "")
    if isinstance(raw_date, datetime.datetime):
        invoice_date = raw_date.strftime("%Y-%m-%d")
    else:
        invoice_date = str(raw_date).strip()

    invoice_json = OrderedDict([
        ("invoiceType", safe(section_data.get("invoiceType"))),
        ("invoiceDate", invoice_date),
        ("sellerNTNCNIC", str(safe(section_data.get("sellerNTNCNIC")))),
        ("sellerBusinessName", safe(section_data.get("sellerBusinessName"))),
        ("sellerProvince", safe(section_data.get("sellerProvince"))),
        ("sellerAddress", safe(section_data.get("sellerAddress"))),
        ("buyerNTNCNIC", str(safe(section_data.get("buyerNTNCNIC")))),
        ("buyerBusinessName", safe(section_data.get("buyerBusinessName"))),
        ("buyerProvince", safe(section_data.get("buyerProvince"))),
        ("buyerAddress", safe(section_data.get("buyerAddress"))),
        ("buyerRegistrationType", safe(section_data.get("buyerRegistrationType"))),
        ("invoiceRefNo", str(safe(section_data.get("invoiceRefNo", "")))),  # critical fix
        ("scenarioId", safe(section_data.get("scenarioId"))),
        ("items", items)
    ])

    last_json_data[env] = invoice_json
    return app.response_class(
        response=json.dumps(invoice_json, indent=2, allow_nan=False),
        mimetype='application/json'
    )










# Submit JSON to FBR
@app.route('/submit-fbr', methods=['POST'])
def submit_fbr():
    env = get_env()
    records = load_records(env)
    if env not in last_json_data:
        return jsonify({'error': 'No JSON to submit'}), 400

    api_token = get_api_token(env)
    api_url = get_api_url(env)
    headers = {
        "Authorization": f"Bearer {api_token}",
        "Content-Type": "application/json"
    }

    try:
        response = requests.post(api_url, headers=headers, json=last_json_data[env])
        res_json = response.json()

        invoice_no = res_json.get("invoiceNumber", "N/A")
        last_json_data[env]["fbrInvoiceNumber"] = invoice_no
        is_success = bool(invoice_no and invoice_no != "N/A")
        status = "Success" if is_success else "Failed"
        date = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        invoice_date = last_json_data[env].get("invoiceDate", "")
        item = last_json_data[env]["items"][0]
        value_sales_ex_st = float(item.get("valueSalesExcludingST", 0))
        sales_tax_applicable = float(item.get("salesTaxApplicable", 0))
        total_value = value_sales_ex_st + sales_tax_applicable

        record = {
            "sr": len(records) + 1,
            "invoiceReference": invoice_no,
            "invoiceType": last_json_data[env]["invoiceType"],
            "invoiceDate": invoice_date,
            "buyerName": last_json_data[env]["buyerBusinessName"],
            "sellerName": last_json_data[env]["sellerBusinessName"],
            "totalValue": total_value,
            "valueSalesExcludingST": value_sales_ex_st,
            "salesTaxApplicable": sales_tax_applicable,
            "status": status,
            "date": date,
            "items": last_json_data[env]["items"]
        }
        records.append(record)
        save_records(env, records)

        return jsonify({
            "status": status,
            "invoiceNumber": invoice_no,
            "date": date
        })

    except Exception as e:
        return jsonify({"error": str(e)}), 500

# Get all past records
@app.route('/records', methods=['GET'])
def get_records():
    env = get_env()
    records = load_records(env)
    return jsonify(records)

@app.route('/')
def index():
    return render_template('index.html')

@app.route('/dashboard.html')
def dashboard_html():
    return render_template('dashboard.html')








def copy_row_format(ws, source_row, target_row):
    for col in range(1, ws.max_column + 1):
        source_cell = ws.cell(row=source_row, column=col)
        target_cell = ws.cell(row=target_row, column=col)

        if isinstance(target_cell, MergedCell):
            continue  # Skip writing styles to merged cells (avoids crash)

        if source_cell.has_style:
            target_cell.font = copy(source_cell.font)
            target_cell.border = copy(source_cell.border)
            target_cell.fill = copy(source_cell.fill)
            target_cell.number_format = copy(source_cell.number_format)
            target_cell.protection = copy(source_cell.protection)
            target_cell.alignment = copy(source_cell.alignment)








@app.route('/generate-invoice-excel', methods=['GET'])
def generate_invoice_excel():
    env = get_env()
    if env not in last_json_data:
        return jsonify({'error': 'No JSON data to generate invoice'}), 400

    data = last_json_data[env]
    items = data['items']
    template_path = 'template.xlsx'

    if not os.path.exists(template_path):
        return jsonify({'error': 'Template file not found'}), 404

    wb = load_workbook(template_path)
    ws = wb.active

    # ------------------- Static Info Mapping -------------------
    ws['C3'] = data.get('sellerBusinessName', '')
    ws['C4'] = data.get('sellerAddress', '')
    ws['C6'] = data.get('sellerProvince', '')
    ws['C7'] = data.get('sellerNTNCNIC', '')

    ws['C9'] = data.get('buyerBusinessName', '')
    ws['C10'] = data.get('buyerAddress', '')
    ws['C11'] = data.get('buyerProvince', '')
    ws['C12'] = data.get('buyerNTNCNIC', '')

    ws['G5'] = data.get('invoiceRefNo', 'N/A')
    try:
        date_obj = datetime.datetime.strptime(data['invoiceDate'], "%Y-%m-%d")
        ws['G6'] = date_obj.strftime("%d.%m.%Y")
    except:
        ws['G6'] = data['invoiceDate']

    # ------------------- Write Line Items Without Spacing -------------------
    start_row = 17
    total_excl = 0
    total_tax = 0

    for i, item in enumerate(items):
        row = start_row + i

        # Copy format from row 17
        if i > 0:
            copy_row_format(ws, 17, row)

        ws[f'A{row}'] = i + 1
        ws[f'B{row}'] = item.get('quantity', '')
        ws[f'C{row}'] = item.get('productDescription', '')
        ws[f'D{row}'] = item.get('hsCode', '')

        try:
            rate = float(str(item.get('rate', '0')).strip('%'))
        except:
            rate = 0
        ws[f'E{row}'] = rate

        try:
            excl = float(str(item.get('valueSalesExcludingST', 0)).replace(",", ""))
        except:
            excl = 0

        try:
            tax = float(str(item.get('salesTaxApplicable', 0)).replace(",", ""))
        except:
            tax = 0

        ws[f'F{row}'] = excl
        ws[f'G{row}'] = tax
        ws[f'H{row}'] = excl + tax

        total_excl += excl
        total_tax += tax

    # ------------------- Custom TOTAL Row (Manually Appended) -------------------
    total_row = start_row + len(items)

    # Apply same style from previous row (optional)
    copy_row_format(ws, 17, total_row)

    ws[f'E{total_row}'] = "TOTAL"
    ws[f'F{total_row}'] = total_excl
    ws[f'G{total_row}'] = total_tax
    ws[f'H{total_row}'] = total_excl + total_tax

    # ------------------- Footer Values (Optional) -------------------
    ws['D25'] = total_tax
    ws['D26'] = total_excl + total_tax
    
    
    
    
    
    
    # --- Add FBR Invoice Number & QR Code ---
    fbr_invoice = data.get("fbrInvoiceNumber")
    if fbr_invoice:
        ws['G8'] = fbr_invoice
        ws['G8'].font = Font(bold=True, size=10)

    # Generate QR code
    qr = qrcode.make(fbr_invoice)
    with tempfile.NamedTemporaryFile(suffix=".png", delete=False) as tmp:
        qr_path = tmp.name
        qr.save(qr_path)

    # Insert into Excel at E11
    img = ExcelImage(qr_path)
    img.width = 100  # adjust size as needed
    img.height = 100
    ws.add_image(img, 'G10')
    
    
    
    
    # Insert FBR logo image
    logo_path = "fbr_logo.png"  # Change path if stored elsewhere
    if os.path.exists(logo_path):
        img = OpenPyxlImage(logo_path)
        img.width = 100   # Resize width (adjust as needed)
        img.height = 90  # Resize height (adjust as needed)
        ws.add_image(img, 'F10')  # Place at cell F10
    
    
    

    # ------------------- Save and Return -------------------
    
    
    output_excel = f'generated_invoice_{env}.xlsx'
    wb.save(output_excel)
    wb.close()

     # Create HTML string from template
    rendered = render_template("invoice_template.html", data=data)
    output_pdf = f'generated_invoice_{env}.pdf'
    pdfkit.from_string(rendered, output_pdf)

    return send_from_directory(directory='.', path=output_pdf, as_attachment=True)
    
    
    
    # output_excel = f'generated_invoice_{env}.xlsx'
    # output_pdf = f'generated_invoice_{env}.pdf'

    # wb.save(output_excel)
    # wb.close()  # Close openpyxl workbook before opening it in Excel

    # # Convert to PDF using Excel
    # pythoncom.CoInitialize()  # Init COM
    
    # excel = win32.gencache.EnsureDispatch('Excel.Application')
    # excel.Visible = False
    # wb_pdf = excel.Workbooks.Open(os.path.abspath(output_excel))
    # wb_pdf.ExportAsFixedFormat(0, os.path.abspath(output_pdf))
    # wb_pdf.Close(False)
    # excel.Quit()
    
    # pythoncom.CoUninitialize()  # Uninit COM

    # # Return the PDF file
    # return send_from_directory(directory='.', path=output_pdf, as_attachment=True)

if __name__ == '__main__':
    app.run(debug=True)
